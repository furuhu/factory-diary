# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from PIL import Image as PILImage, ImageOps
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import os
from io import BytesIO
import math

# PDF Generation Libraries
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib import units
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
# Import CJK Font support
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

# --- Streamlit UI 設定 ---
st.set_page_config(page_title="工廠安裝日記", layout="wide")

# --- Try to Register CJK Font ---
try:
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    CJK_FONT_NAME = 'STSong-Light'
except Exception as e:
    CJK_FONT_NAME = 'Helvetica' # Fallback font

# --- Streamlit 應用程式標題 ---
st.title("🛠️ 工廠安裝日記自動生成器")
if CJK_FONT_NAME == 'STSong-Light':
    st.success("中文字體 (STSong-Light) 加載成功，可用於 PDF 導出。")
else:
    st.warning(f"無法加載中文字體 STSong-Light，PDF 中的中文可能無法正確顯示。將使用 {CJK_FONT_NAME}。")

# --- 新增：報告標題 ---
st.header("📝 報告標題")
report_title_input = st.text_input("請輸入報告主標題 (例如：XX專案安裝日記 - YYY設備)")


# --- 基本資料欄位 ---
st.header("📅 基本資訊")
col1, col3 = st.columns(2) # 移除天氣欄位後，改為2欄
with col1:
    install_date = st.date_input("安裝日期 (將作為新分頁名稱)", value=date.today())
# with col2: # 天氣欄位已刪除
    # weather_options = ["晴", "陰", "多雲", "陣雨", "雷陣雨", "小雨", "大雨", "其他"]
    # weather = st.selectbox("天氣", options=weather_options, index=0)
with col3:
    recorder = st.text_input("記錄人")

# --- 新增：參加人員 ---
st.header("🧑‍🤝‍🧑 參加人員")
attendees = st.text_area("請輸入參加人員 (每行一位，或用逗號分隔)", height=100)


# --- 人力配置 ---
st.header("👥 人力配置")
st.write("請填寫供應商人員與外包人員的分類人數")
# ******** 修改：土木 -> 業務 ********
role_types = ["機械", "電機", "業務", "軟體"]
# ***********************************
staff_data = {}
# 供應商人員輸入
cols_sup = st.columns(len(role_types) + 1)
cols_sup[0].markdown("#### 供應商人員")
staff_data['供應商人員'] = []
for i, role in enumerate(role_types):
    count = cols_sup[i+1].number_input(f"供應商-{role}", min_value=0, step=1, key=f"sup_{role}")
    staff_data['供應商人員'].append(count)
# 外包人員輸入
cols_sub = st.columns(len(role_types) + 1)
cols_sub[0].markdown("#### 外包人員")
staff_data['外包人員'] = []
for i, role in enumerate(role_types):
    count = cols_sub[i+1].number_input(f"外包-{role}", min_value=0, step=1, key=f"sub_{role}")
    staff_data['外包人員'].append(count)

# --- 裝機進度 ---
if "machine_sections" not in st.session_state: st.session_state["machine_sections"] = []
st.header("🏗️ 裝機進度紀錄")
new_machine_name = st.text_input("輸入新機台名稱", key="new_machine_input")
add_machine_button = st.button("➕ 新增機台")
if add_machine_button and new_machine_name:
    if new_machine_name not in st.session_state["machine_sections"]:
        st.session_state["machine_sections"].append(new_machine_name); st.success(f"已新增機台: {new_machine_name}")
progress_entries = []
for idx, machine_name in enumerate(st.session_state["machine_sections"]):
    with st.expander(f"🔧 {machine_name} (點此展開/收合)", expanded=True):
        for i in range(1, 5): # 裝機進度維持4項
            st.markdown(f"**第 {i} 項**"); cols = st.columns([4, 1, 2])
            content = cols[0].text_input(f"內容", key=f"machine_{idx}_content_{i}")
            manpower = cols[1].number_input(f"人力", key=f"machine_{idx}_manpower_{i}", min_value=0, step=1)
            note = cols[2].text_input(f"備註", key=f"machine_{idx}_note_{i}")
            if content: progress_entries.append([machine_name, i, content, manpower, note])

# --- 週邊工作 ---
st.header("🔧 週邊工作紀錄")
# ******** 修改：增加到 10 項 ********
side_entries = []
for i in range(1, 11): # 項目 1 到 10
# ***********************************
    st.markdown(f"**第 {i} 項**"); cols = st.columns([4, 1, 2])
    content = cols[0].text_input(f"內容 ", key=f"side_content_{i}")
    manpower = cols[1].number_input(f"人力 ", key=f"side_manpower_{i}", min_value=0, step=1)
    note = cols[2].text_input(f"備註 ", key=f"side_note_{i}")
    if content: side_entries.append([i, content, manpower, note])

# --- 照片上傳 ---
st.header("📸 上傳照片")
st.markdown("**進度留影**")
photos = st.file_uploader("上傳今天的照片（jpg/png/jpeg）", type=["jpg", "jpeg", "png"], accept_multiple_files=True, key="photo_uploader")

# --- 新增：上傳舊 Excel 檔案 ---
st.header("📂 合併舊日誌 (可選)")
uploaded_excel_file = st.file_uploader("上傳之前的 Excel 安裝日記檔案 (若要合併)", type=["xlsx"])

# --- 導出按鈕 ---
st.header("📄 導出報告")
col_export1, col_export2 = st.columns(2)

# --- 輔助函數：定義 Excel 樣式 (移到按鈕外部) ---
bold_font_excel = Font(name="標楷體", size=11, bold=True)
normal_font_excel = Font(name="標楷體", size=11)
title_font_excel = Font(name="標楷體", size=14, bold=True) # 新增報告標題字體
thin_border_side_excel = Side(style='thin', color='000000')
thin_border_excel = Border(left=thin_border_side_excel, right=thin_border_side_excel, top=thin_border_side_excel, bottom=thin_border_side_excel)
center_align_wrap_excel = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align_wrap_excel = Alignment(horizontal="left", vertical="center", wrap_text=True)
DEFAULT_COL_WIDTH_EXCEL = 18
DEFAULT_ROW_HEIGHT_EXCEL = 25
IMAGE_ROW_HEIGHT_EXCEL = 120
NUM_COLS_TOTAL_EXCEL = 6

# --- 輔助函數：將一天資料寫入指定的 Excel 工作表 ---
def write_day_to_excel_sheet(ws, report_title_ws, install_date_ws, attendees_ws, recorder_ws, staff_data_ws, progress_entries_ws, side_entries_ws, photos_ws):
    """將一天的所有資料寫入指定的 openpyxl worksheet (ws)"""

    for i in range(1, NUM_COLS_TOTAL_EXCEL + 1):
        ws.column_dimensions[get_column_letter(i)].width = DEFAULT_COL_WIDTH_EXCEL
    current_row_ws = 1

    def write_styled_cell_internal(row, col, value, font, alignment, border=thin_border_excel):
        cell = ws.cell(row=row, column=col)
        cell.value = value; cell.font = font; cell.alignment = alignment
        if border: cell.border = border
        current_height = ws.row_dimensions[row].height
        if current_height is None or current_height < DEFAULT_ROW_HEIGHT_EXCEL: ws.row_dimensions[row].height = DEFAULT_ROW_HEIGHT_EXCEL

    def apply_styles_only_internal(row, col, font, alignment, border=thin_border_excel):
        cell = ws.cell(row=row, column=col)
        cell.font = font; cell.alignment = alignment
        if border: cell.border = border
        current_height = ws.row_dimensions[row].height
        if current_height is None or current_height < DEFAULT_ROW_HEIGHT_EXCEL: ws.row_dimensions[row].height = DEFAULT_ROW_HEIGHT_EXCEL

    # --- 新增：寫入報告標題 ---
    if report_title_ws:
        ws.merge_cells(start_row=current_row_ws, start_column=1, end_row=current_row_ws, end_column=NUM_COLS_TOTAL_EXCEL)
        write_styled_cell_internal(current_row_ws, 1, report_title_ws, title_font_excel, center_align_wrap_excel, border=None)
        ws.row_dimensions[current_row_ws].height = 30 # 加高標題列
        current_row_ws += 1
        ws.row_dimensions[current_row_ws].height = DEFAULT_ROW_HEIGHT_EXCEL # 空行
        current_row_ws += 1


    # --- 寫入基本資訊 ---
    ws.merge_cells(start_row=current_row_ws, start_column=2, end_row=current_row_ws, end_column=NUM_COLS_TOTAL_EXCEL)
    write_styled_cell_internal(current_row_ws, 1, "日期", bold_font_excel, center_align_wrap_excel)
    write_styled_cell_internal(current_row_ws, 2, str(install_date_ws), normal_font_excel, center_align_wrap_excel)
    for c in range(3, NUM_COLS_TOTAL_EXCEL + 1): apply_styles_only_internal(current_row_ws, c, normal_font_excel, center_align_wrap_excel, thin_border_excel)
    current_row_ws += 1
    # 天氣欄位已刪除

    # --- 新增：寫入參加人員 ---
    if attendees_ws:
        ws.merge_cells(start_row=current_row_ws, start_column=2, end_row=current_row_ws, end_column=NUM_COLS_TOTAL_EXCEL)
        write_styled_cell_internal(current_row_ws, 1, "參加人員", bold_font_excel, left_align_wrap_excel) # 靠左
        write_styled_cell_internal(current_row_ws, 2, attendees_ws, normal_font_excel, left_align_wrap_excel) # 靠左
        for c in range(3, NUM_COLS_TOTAL_EXCEL + 1): apply_styles_only_internal(current_row_ws, c, normal_font_excel, left_align_wrap_excel, thin_border_excel)
        current_row_ws += 1

    ws.row_dimensions[current_row_ws].height = DEFAULT_ROW_HEIGHT_EXCEL; current_row_ws += 1

    # --- 寫入人力配置 ---
    header_staff = ["人員分類", *role_types, "總計"] # role_types 已更新
    for col_idx, header_text in enumerate(header_staff, 1):
        if col_idx <= NUM_COLS_TOTAL_EXCEL: write_styled_cell_internal(current_row_ws, col_idx, header_text, bold_font_excel, center_align_wrap_excel)
    current_row_ws += 1
    for group in ["供應商人員", "外包人員"]:
        group_counts = staff_data_ws.get(group, [])
        processed_counts = []; valid_data = True
        if isinstance(group_counts, list):
            for item in group_counts:
                if isinstance(item, (int, float)): processed_counts.append(item)
                else:
                    try: processed_counts.append(int(item))
                    except (ValueError, TypeError): valid_data = False; processed_counts.append(0)
        else: valid_data = False; processed_counts = [0] * len(role_types)
        total = sum(processed_counts) if valid_data or processed_counts else 0
        row_data = [group, *processed_counts, total]
        for col_idx, cell_value in enumerate(row_data, 1):
             if col_idx <= NUM_COLS_TOTAL_EXCEL:
                align = left_align_wrap_excel if col_idx == 1 else center_align_wrap_excel
                write_styled_cell_internal(current_row_ws, col_idx, cell_value, normal_font_excel, align)
        current_row_ws += 1
    ws.row_dimensions[current_row_ws].height = DEFAULT_ROW_HEIGHT_EXCEL; current_row_ws += 1

    # --- 寫入裝機進度 ---
    if progress_entries_ws:
        ws.merge_cells(start_row=current_row_ws, start_column=1, end_row=current_row_ws, end_column=NUM_COLS_TOTAL_EXCEL)
        write_styled_cell_internal(current_row_ws, 1, "裝機進度", bold_font_excel, center_align_wrap_excel)
        for c in range(2, NUM_COLS_TOTAL_EXCEL + 1): apply_styles_only_internal(current_row_ws, c, bold_font_excel, center_align_wrap_excel, thin_border_excel)
        current_row_ws += 1
        header_progress = ["機台", "項次", "內容", "人力", "備註"]
        write_styled_cell_internal(current_row_ws, 1, header_progress[0], bold_font_excel, center_align_wrap_excel); write_styled_cell_internal(current_row_ws, 2, header_progress[1], bold_font_excel, center_align_wrap_excel)
        ws.merge_cells(start_row=current_row_ws, start_column=3, end_row=current_row_ws, end_column=4); write_styled_cell_internal(current_row_ws, 3, header_progress[2], bold_font_excel, center_align_wrap_excel)
        apply_styles_only_internal(current_row_ws, 4, bold_font_excel, center_align_wrap_excel, thin_border_excel); write_styled_cell_internal(current_row_ws, 5, header_progress[3], bold_font_excel, center_align_wrap_excel); write_styled_cell_internal(current_row_ws, 6, header_progress[4], bold_font_excel, center_align_wrap_excel)
        current_row_ws += 1
        for row_data in progress_entries_ws:
            machine, item, content, manpower, note = row_data
            write_styled_cell_internal(current_row_ws, 1, machine, normal_font_excel, left_align_wrap_excel); write_styled_cell_internal(current_row_ws, 2, item, normal_font_excel, center_align_wrap_excel)
            ws.merge_cells(start_row=current_row_ws, start_column=3, end_row=current_row_ws, end_column=4); write_styled_cell_internal(current_row_ws, 3, content, normal_font_excel, left_align_wrap_excel) # 內容靠左
            apply_styles_only_internal(current_row_ws, 4, normal_font_excel, left_align_wrap_excel, thin_border_excel); write_styled_cell_internal(current_row_ws, 5, manpower, normal_font_excel, center_align_wrap_excel); write_styled_cell_internal(current_row_ws, 6, note, normal_font_excel, left_align_wrap_excel) # 備註靠左
            current_row_ws += 1
        ws.row_dimensions[current_row_ws].height = DEFAULT_ROW_HEIGHT_EXCEL; current_row_ws += 1

    # --- 寫入週邊工作 ---
    if side_entries_ws: # 週邊工作項數已在 UI 增加
        ws.merge_cells(start_row=current_row_ws, start_column=1, end_row=current_row_ws, end_column=NUM_COLS_TOTAL_EXCEL)
        write_styled_cell_internal(current_row_ws, 1, "週邊工作", bold_font_excel, center_align_wrap_excel)
        for c in range(2, NUM_COLS_TOTAL_EXCEL + 1): apply_styles_only_internal(current_row_ws, c, bold_font_excel, center_align_wrap_excel, thin_border_excel)
        current_row_ws += 1
        header_side = ["項次", "內容", "人力", "備註"]
        write_styled_cell_internal(current_row_ws, 1, header_side[0], bold_font_excel, center_align_wrap_excel)
        ws.merge_cells(start_row=current_row_ws, start_column=2, end_row=current_row_ws, end_column=4); write_styled_cell_internal(current_row_ws, 2, header_side[1], bold_font_excel, center_align_wrap_excel)
        apply_styles_only_internal(current_row_ws, 3, bold_font_excel, center_align_wrap_excel, thin_border_excel); apply_styles_only_internal(current_row_ws, 4, bold_font_excel, center_align_wrap_excel, thin_border_excel)
        write_styled_cell_internal(current_row_ws, 5, header_side[2], bold_font_excel, center_align_wrap_excel); write_styled_cell_internal(current_row_ws, 6, header_side[3], bold_font_excel, center_align_wrap_excel)
        current_row_ws += 1
        for row_data in side_entries_ws:
            item, content, manpower, note = row_data
            write_styled_cell_internal(current_row_ws, 1, item, normal_font_excel, center_align_wrap_excel)
            ws.merge_cells(start_row=current_row_ws, start_column=2, end_row=current_row_ws, end_column=4); write_styled_cell_internal(current_row_ws, 2, content, normal_font_excel, left_align_wrap_excel) # 內容靠左
            apply_styles_only_internal(current_row_ws, 3, normal_font_excel, left_align_wrap_excel, thin_border_excel); apply_styles_only_internal(current_row_ws, 4, normal_font_excel, left_align_wrap_excel, thin_border_excel)
            write_styled_cell_internal(current_row_ws, 5, manpower, normal_font_excel, center_align_wrap_excel); write_styled_cell_internal(current_row_ws, 6, note, normal_font_excel, left_align_wrap_excel) # 備註靠左
            current_row_ws += 1
        ws.row_dimensions[current_row_ws].height = DEFAULT_ROW_HEIGHT_EXCEL; current_row_ws += 1

    # --- 處理圖片區域 ---
    if photos_ws:
        ws.row_dimensions[current_row_ws].height = DEFAULT_ROW_HEIGHT_EXCEL; current_row_ws += 1
        ws.merge_cells(start_row=current_row_ws, start_column=1, end_row=current_row_ws, end_column=NUM_COLS_TOTAL_EXCEL)
        write_styled_cell_internal(current_row_ws, 1, "進度留影", bold_font_excel, center_align_wrap_excel, border=None)
        for c in range(2, NUM_COLS_TOTAL_EXCEL + 1): apply_styles_only_internal(current_row_ws, c, bold_font_excel, center_align_wrap_excel, border=None)
        ws.row_dimensions[current_row_ws].height = DEFAULT_ROW_HEIGHT_EXCEL; current_row_ws += 1
        try: default_char_width_approx = 7; target_img_width_px = int(DEFAULT_COL_WIDTH_EXCEL * 3 * default_char_width_approx)
        except: target_img_width_px = int(18 * 3 * 7)
        target_img_height_px = int(IMAGE_ROW_HEIGHT_EXCEL / 0.75)
        width_adjustment = 8; adjusted_target_width_px = max(1, target_img_width_px - width_adjustment)
        img_col_width = 3; num_img_cols = 2
        for i in range(0, len(photos_ws), num_img_cols):
            ws.row_dimensions[current_row_ws].height = IMAGE_ROW_HEIGHT_EXCEL
            ws.row_dimensions[current_row_ws + 1].height = DEFAULT_ROW_HEIGHT_EXCEL
            for j in range(num_img_cols):
                photo_index = i + j
                if photo_index < len(photos_ws):
                    img_file = photos_ws[photo_index]; filename = img_file.name
                    try:
                        img = PILImage.open(img_file); img = ImageOps.exif_transpose(img)
                        img_w, img_h = img.size; assert img_w > 0 and img_h > 0
                        target_size = (adjusted_target_width_px, target_img_height_px)
                        img_cropped = ImageOps.fit(img, target_size, method=PILImage.Resampling.LANCZOS)
                        img_buffer = BytesIO(); img_cropped.save(img_buffer, format='PNG'); img_buffer.seek(0)
                        col_start = 1 + j * img_col_width; anchor_cell = f"{get_column_letter(col_start)}{current_row_ws}"
                        xl_img = XLImage(img_buffer); ws.add_image(xl_img, anchor_cell)
                        col_end = col_start + img_col_width - 1
                        merge_range_caption = f"{get_column_letter(col_start)}{current_row_ws + 1}:{get_column_letter(col_end)}{current_row_ws + 1}"
                        ws.merge_cells(merge_range_caption)
                        write_styled_cell_internal(current_row_ws + 1, col_start, f"說明：{filename}", normal_font_excel, center_align_wrap_excel)
                        for c_idx in range(col_start + 1, col_end + 1): apply_styles_only_internal(current_row_ws + 1, c_idx, normal_font_excel, center_align_wrap_excel, thin_border_excel)
                        for r_idx in [current_row_ws]:
                            for c_idx in range(col_start, col_end + 1): apply_styles_only_internal(r_idx, c_idx, normal_font_excel, Alignment(vertical="center"), thin_border_excel)
                    except Exception as e:
                        st.error(f"處理圖片 {filename} 時發生錯誤 (將在 Excel 中標記): {e}")
                        col_start = 1 + j * img_col_width; col_end = col_start + img_col_width - 1
                        merge_range_caption = f"{get_column_letter(col_start)}{current_row_ws + 1}:{get_column_letter(col_end)}{current_row_ws + 1}"
                        try: ws.merge_cells(merge_range_caption)
                        except: pass
                        write_styled_cell_internal(current_row_ws + 1, col_start, f"圖片錯誤", normal_font_excel, center_align_wrap_excel)
                        for c_idx in range(col_start + 1, col_end + 1): apply_styles_only_internal(current_row_ws + 1, c_idx, normal_font_excel, center_align_wrap_excel, thin_border_excel)
            current_row_ws += 2

    # --- 添加記錄人資訊 (Excel 底部) ---
    ws.row_dimensions[current_row_ws].height = DEFAULT_ROW_HEIGHT_EXCEL
    current_row_ws += 1
    recorder_text = f"記錄人： {recorder_ws}"
    merge_start_col = 1; merge_end_col = NUM_COLS_TOTAL_EXCEL
    merge_range_recorder = f"{get_column_letter(merge_start_col)}{current_row_ws}:{get_column_letter(merge_end_col)}{current_row_ws}"
    try: ws.merge_cells(merge_range_recorder)
    except Exception as merge_err: merge_end_col = merge_start_col
    write_styled_cell_internal(current_row_ws, merge_start_col, recorder_text, normal_font_excel, left_align_wrap_excel, border=thin_border_excel)
    if merge_end_col > merge_start_col:
        for c in range(merge_start_col + 1, merge_end_col + 1): apply_styles_only_internal(current_row_ws, c, normal_font_excel, left_align_wrap_excel, border=thin_border_excel)

# --- Excel 導出按鈕邏輯 ---
with col_export1:
    if st.button("✅ 產出/合併 Excel"):
        current_report_title = report_title_input
        current_install_date = install_date
        current_attendees = attendees
        current_recorder = recorder
        current_staff_data = staff_data
        current_progress_entries = progress_entries
        current_side_entries = side_entries
        current_photos = photos
        new_sheet_name = current_install_date.strftime("%Y-%m-%d")

        wb = None
        if uploaded_excel_file is not None:
            try:
                wb = load_workbook(uploaded_excel_file)
                st.info(f"已加載舊檔案: {uploaded_excel_file.name}。將添加/覆蓋分頁 '{new_sheet_name}'。")
                if new_sheet_name in wb.sheetnames:
                    # st.warning(f"工作表 '{new_sheet_name}' 已存在，將覆蓋其內容。")
                    ws = wb[new_sheet_name]
                    # 簡單起見，這裡不清除舊內容，直接覆蓋
                else:
                    ws = wb.create_sheet(title=new_sheet_name)
            except Exception as e:
                st.error(f"讀取上傳的 Excel 檔案時出錯: {e}")
                st.warning("將創建全新的 Excel 檔案。")
                wb = None
        if wb is None:
            wb = Workbook()
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1 : # 移除預設工作表 (如果有多個)
                del wb["Sheet"]
            ws = wb.active
            ws.title = new_sheet_name
            # st.info(f"創建新 Excel 檔案，分頁 '{new_sheet_name}'。")

        try:
            write_day_to_excel_sheet(ws, current_report_title, current_install_date, current_attendees, current_recorder,
                                     current_staff_data, current_progress_entries, current_side_entries, current_photos)
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            excel_file_name = f"{current_report_title}_{current_install_date.strftime('%Y%m%d')}.xlsx" if current_report_title else f"安裝日記_{current_install_date.strftime('%Y%m%d')}.xlsx"
            if uploaded_excel_file and current_report_title: # 如果合併且有報告標題
                 excel_file_name = f"{current_report_title}_合併_{current_install_date.strftime('%Y%m%d')}.xlsx"
            elif uploaded_excel_file: # 如果合併但無報告標題
                 excel_file_name = f"安裝日記_合併_{current_install_date.strftime('%Y%m%d')}.xlsx"


            st.download_button(label="📥 下載 Excel 檔案", data=excel_file, file_name=excel_file_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.success(f"檔案 {excel_file_name} 已成功產生/合併！")
        except Exception as write_err:
             st.error(f"寫入資料到 Excel 工作表 '{ws.title}' 時發生錯誤: {write_err}")

# --- PDF 導出按鈕邏輯 (只產生當天資料) ---
with col_export2:
    if st.button("📄 產出 PDF 報告 (僅當天)"):
        st.info("PDF 報告目前只會包含您在頁面上輸入的當天資料。")
        pdf_buffer = BytesIO()
        page_width, page_height = A4
        margin = 1.5*units.cm
        doc_width = page_width - 2 * margin

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='CJKNormal', parent=styles['Normal'], fontName=CJK_FONT_NAME, fontSize=10, alignment=TA_LEFT))
        styles.add(ParagraphStyle(name='CJKBold', parent=styles['CJKNormal'], fontName=CJK_FONT_NAME, fontSize=10, alignment=TA_LEFT))
        # 主標題 (報告標題)
        styles.add(ParagraphStyle(name='CJKMainTitle', parent=styles['h1'], fontName=CJK_FONT_NAME, fontSize=20, alignment=TA_CENTER, spaceAfter=6))
        # 副標題 (工廠安裝日記)
        styles.add(ParagraphStyle(name='CJKSubTitle', parent=styles['h2'], fontName=CJK_FONT_NAME, fontSize=16, alignment=TA_CENTER, spaceAfter=12))
        styles.add(ParagraphStyle(name='CJKHeading2', fontName=CJK_FONT_NAME, fontSize=14, leading=17, alignment=TA_LEFT, spaceBefore=6, spaceAfter=6))
        styles.add(ParagraphStyle(name='CJKTableContent', parent=styles['Normal'], fontName=CJK_FONT_NAME, fontSize=9, alignment=TA_CENTER))
        styles.add(ParagraphStyle(name='CJKTableContentLeft', parent=styles['CJKTableContent'], alignment=TA_LEFT))
        styles.add(ParagraphStyle(name='CJKFooterTitleBold', fontName=CJK_FONT_NAME, fontSize=14, alignment=TA_LEFT, leading=17))

        doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=margin, rightMargin=margin, topMargin=margin, bottomMargin=margin, title=f"安裝日記_{install_date}", author="工廠安裝日記自動生成器")
        story = []

        # --- PDF 內容 - 第一頁 ---
        if report_title_input:
            story.append(Paragraph(report_title_input, styles['CJKMainTitle']))
            story.append(Paragraph("安裝日誌", styles['CJKSubTitle']))
        else:
            story.append(Paragraph("工廠安裝日記", styles['CJKMainTitle'])) # 如果沒有輸入報告標題，使用預設
        story.append(Spacer(1, 0.5*units.cm))

        # 基本資訊表格
        basic_info_data = [
            [Paragraph("<b>日期</b>", styles['CJKNormal']), Paragraph(str(install_date), styles['CJKNormal'])],
            # 天氣已刪除
        ]
        basic_info_table = Table(basic_info_data, colWidths=[doc_width/4, doc_width*3/4]) # 調整欄寬以適應兩欄
        basic_info_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
        story.append(basic_info_table); story.append(Spacer(1, 0.2*units.cm)) # 縮小間距

        # 參加人員
        if attendees:
            story.append(Paragraph("<b>參加人員：</b>", styles['CJKNormal']))
            # 將 attendees 字串按換行符分割，然後用逗號連接（如果使用者用換行輸入）
            # 或者直接顯示 text_area 的內容
            attendees_display = attendees.replace('\n', ', ')
            story.append(Paragraph(attendees_display, styles['CJKNormal']))
            story.append(Spacer(1, 0.5*units.cm))


        story.append(Paragraph("人力配置", styles['CJKHeading2']))
        staff_header = [Paragraph(f"<b>{h}</b>", styles['CJKTableContent']) for h in ["人員分類", *role_types, "總計"]]
        staff_table_data = [staff_header]
        for group in ["供應商人員", "外包人員"]:
            group_counts = staff_data.get(group, []); processed_counts = [int(c) for c in group_counts]; total = sum(processed_counts)
            row_data_text = [Paragraph(group, styles['CJKTableContentLeft'])] + [Paragraph(str(c), styles['CJKTableContent']) for c in processed_counts] + [Paragraph(str(total), styles['CJKTableContent'])]
            staff_table_data.append(row_data_text)
        staff_col_widths = [doc_width*0.225] + [doc_width*0.15]*len(role_types) + [doc_width*0.175]
        staff_table = Table(staff_table_data, colWidths=staff_col_widths)
        staff_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black), ('BACKGROUND', (0,0), (-1,0), colors.lightgrey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (1,0), (-1,-1), 'CENTER')]))
        story.append(staff_table); story.append(Spacer(1, 0.5*units.cm))

        if progress_entries:
            story.append(Paragraph("裝機進度紀錄", styles['CJKHeading2']))
            progress_header = [Paragraph(f"<b>{h}</b>", styles['CJKTableContent']) for h in ["機台", "項次", "內容", "人力", "備註"]]
            progress_table_data = [progress_header]
            for entry in progress_entries:
                row_data_text = [Paragraph(str(entry[0]), styles['CJKTableContentLeft']), Paragraph(str(entry[1]), styles['CJKTableContent']), Paragraph(str(entry[2]), styles['CJKTableContentLeft']), Paragraph(str(entry[3]), styles['CJKTableContent']), Paragraph(str(entry[4]), styles['CJKTableContentLeft'])]
                progress_table_data.append(row_data_text)
            progress_table = Table(progress_table_data, colWidths=[doc_width*0.15, doc_width*0.1, doc_width*0.4, doc_width*0.1, doc_width*0.25])
            progress_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black), ('BACKGROUND', (0,0), (-1,0), colors.lightgrey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (1,1), (1,-1), 'CENTER'), ('ALIGN', (3,1), (3,-1), 'CENTER')]))
            story.append(progress_table); story.append(Spacer(1, 0.5*units.cm))

        if side_entries: # 週邊工作項數已增加
            story.append(Paragraph("週邊工作紀錄", styles['CJKHeading2']))
            side_header = [Paragraph(f"<b>{h}</b>", styles['CJKTableContent']) for h in ["項次", "內容", "人力", "備註"]]
            side_table_data = [side_header]
            for entry in side_entries:
                row_data_text = [Paragraph(str(entry[0]), styles['CJKTableContent']), Paragraph(str(entry[1]), styles['CJKTableContentLeft']), Paragraph(str(entry[2]), styles['CJKTableContent']), Paragraph(str(entry[3]), styles['CJKTableContentLeft'])]
                side_table_data.append(row_data_text)
            side_table = Table(side_table_data, colWidths=[doc_width*0.1, doc_width*0.55, doc_width*0.1, doc_width*0.25])
            side_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black), ('BACKGROUND', (0,0), (-1,0), colors.lightgrey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (0,1), (0,-1), 'CENTER'), ('ALIGN', (2,1), (2,-1), 'CENTER')]))
            story.append(side_table); story.append(Spacer(1, 0.5*units.cm))

        # --- 換頁 ---
        story.append(PageBreak())

        # --- PDF 內容 - 第二頁 (圖片) ---
        story.append(Paragraph("進度留影", styles['CJKHeading2']))
        story.append(Spacer(1, 0.5*units.cm))

        if photos:
            img_margin = 0.5 * units.cm
            img_width_pt = (doc_width - img_margin) / 2
            img_height_pt = 6 * units.cm
            target_width_px = int(img_width_pt * (4/3))
            target_height_px = int(img_height_pt * (4/3))
            target_size_px = (target_width_px, target_height_px)

            for i in range(0, len(photos), 2):
                img_row_content = []
                img_file_left = photos[i]
                try:
                    img_pil_left = PILImage.open(img_file_left); img_pil_left = ImageOps.exif_transpose(img_pil_left)
                    img_cropped_left = ImageOps.fit(img_pil_left, target_size_px, method=PILImage.Resampling.LANCZOS)
                    img_buffer_left = BytesIO(); img_cropped_left.save(img_buffer_left, format='PNG'); img_buffer_left.seek(0)
                    rl_img_left = Image(img_buffer_left, width=img_width_pt, height=img_height_pt)
                    img_row_content.append(rl_img_left)
                except Exception as img_err:
                    st.error(f"處理圖片 {img_file_left.name} 時發生錯誤: {img_err}")
                    img_row_content.append(Paragraph(f"[圖片錯誤: {img_file_left.name}]", styles['CJKNormal']))

                img_row_content.append(Spacer(img_margin, 1))

                if i + 1 < len(photos):
                    img_file_right = photos[i+1]
                    try:
                        img_pil_right = PILImage.open(img_file_right); img_pil_right = ImageOps.exif_transpose(img_pil_right)
                        img_cropped_right = ImageOps.fit(img_pil_right, target_size_px, method=PILImage.Resampling.LANCZOS)
                        img_buffer_right = BytesIO(); img_cropped_right.save(img_buffer_right, format='PNG'); img_buffer_right.seek(0)
                        rl_img_right = Image(img_buffer_right, width=img_width_pt, height=img_height_pt)
                        img_row_content.append(rl_img_right)
                    except Exception as img_err:
                        st.error(f"處理圖片 {img_file_right.name} 時發生錯誤: {img_err}")
                        img_row_content.append(Paragraph(f"[圖片錯誤: {img_file_right.name}]", styles['CJKNormal']))
                else:
                    img_row_content.append(Spacer(img_width_pt, img_height_pt))

                img_table = Table([img_row_content], colWidths=[img_width_pt, img_margin, img_width_pt])
                img_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
                story.append(img_table)
                story.append(Spacer(1, 0.5*units.cm))

        # --- PDF 內容 - 結尾記錄人 ---
        story.append(Spacer(1, 1*units.cm))
        story.append(Paragraph(f"<b>記錄人： {recorder}</b>", styles['CJKFooterTitleBold']))

        # --- 生成 PDF ---
        try:
            doc.build(story)
            st.success("PDF 報告已成功產生！")
            pdf_buffer.seek(0)
            pdf_file_name = f"{report_title_input}_{install_date}.pdf" if report_title_input else f"安裝日記_{install_date}.pdf"
            st.download_button(label="📥 下載 PDF 報告", data=pdf_buffer, file_name=pdf_file_name, mime="application/pdf")
        except Exception as pdf_err:
            st.error(f"產生 PDF 時發生錯誤: {pdf_err}")
            st.error("可能的原因包括：中文字體問題、圖片處理錯誤或 ReportLab 內部錯誤。請檢查 Streamlit 終端輸出獲取更詳細的錯誤信息。")

# --- Script End ---
